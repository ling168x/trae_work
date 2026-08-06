"""
消息导出模块
支持导出为 CSV, HTML, TXT, XLSX 格式
"""

import os
import csv
import json
import html
from datetime import datetime


# 微信消息类型映射
MSG_TYPE_MAP = {
    1: '文本',
    3: '图片',
    34: '语音',
    43: '视频',
    47: '表情',
    49: '链接/文件/小程序',
    10000: '系统消息',
    10002: '系统消息',
    436207665: '红包',
}


def get_msg_type_name(msg_type):
    """获取消息类型名称"""
    return MSG_TYPE_MAP.get(msg_type, f'未知({msg_type})')


def format_timestamp(ts):
    """格式化微信时间戳 (微信时间戳基于 1970-01-01)"""
    if not ts:
        return ''
    try:
        dt = datetime.fromtimestamp(int(ts))
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return str(ts)


def get_content(msg):
    """智能提取消息内容"""
    # 优先使用 StrContent
    content = msg.get('StrContent', '') or msg.get('strcontent', '')
    if content:
        return content

    # 如果类型是 1 (文本), 使用 Content
    if msg.get('Type', 0) == 1:
        return msg.get('Content', '') or msg.get('content', '') or ''

    # 其他类型返回类型描述
    return f"[{get_msg_type_name(msg.get('Type', 0))}]"


def get_talker_name(msg, contacts_map):
    """获取发送者昵称"""
    talker = msg.get('Talker', '') or msg.get('talker', '')
    if not talker:
        return '未知'

    if talker in contacts_map:
        c = contacts_map[talker]
        return c.get('remark') or c.get('nickname') or talker

    return talker


def get_sender_name(msg, contacts_map):
    """获取消息发送者 (群聊中是谁发的)"""
    is_sender = msg.get('IsSender', 0) or msg.get('issender', 0)
    if is_sender:
        return '我'

    sender = msg.get('Sender', '') or msg.get('sender', '')
    if not sender:
        return '未知'

    if sender in contacts_map:
        c = contacts_map[sender]
        return c.get('remark') or c.get('nickname') or sender

    # 尝试从 Content 中提取发送者 (群聊消息格式: "sender:\ncontent")
    content = msg.get('Content', '') or msg.get('content', '')
    if content and ':\n' in content:
        possible_sender = content.split(':\n')[0]
        if possible_sender and '@chatroom' not in possible_sender:
            return possible_sender

    return sender


def build_contacts_map(contacts):
    """构建联系人映射"""
    return {c['username']: c for c in contacts}


def export_to_csv(messages, output_path, contacts_map=None):
    """
    导出消息到 CSV 文件
    """
    if contacts_map is None:
        contacts_map = {}

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['时间', '发送者', '消息类型', '内容', '原始数据'])

        for msg in messages:
            time_str = format_timestamp(
                msg.get('CreateTime') or msg.get('createtime')
            )
            sender = get_sender_name(msg, contacts_map)
            msg_type = get_msg_type_name(msg.get('Type', 0))
            content = get_content(msg)
            raw = json.dumps(msg, ensure_ascii=False)

            writer.writerow([time_str, sender, msg_type, content, raw])

    print(f"[+] 已导出 {len(messages)} 条消息到: {output_path}")


def export_to_txt(messages, output_path, contacts_map=None):
    """
    导出消息到 TXT 纯文本文件
    """
    if contacts_map is None:
        contacts_map = {}

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("微信聊天记录导出\n")
        f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"消息数量: {len(messages)}\n")
        f.write("=" * 60 + "\n\n")

        for msg in messages:
            time_str = format_timestamp(
                msg.get('CreateTime') or msg.get('createtime')
            )
            sender = get_sender_name(msg, contacts_map)
            msg_type = get_msg_type_name(msg.get('Type', 0))
            content = get_content(msg)

            f.write(f"[{time_str}] {sender} ({msg_type}):\n")
            f.write(f"  {content}\n")
            f.write("-" * 40 + "\n")

    print(f"[+] 已导出 {len(messages)} 条消息到: {output_path}")


def export_to_html(messages, output_path, contacts_map=None, title="微信聊天记录"):
    """
    导出消息到 HTML 文件 (美观的网页格式)
    """
    if contacts_map is None:
        contacts_map = {}

    rows_html = []
    for msg in messages:
        time_str = format_timestamp(
            msg.get('CreateTime') or msg.get('createtime')
        )
        sender = get_sender_name(msg, contacts_map)
        msg_type = get_msg_type_name(msg.get('Type', 0))
        content = html.escape(get_content(msg))

        is_me = 'me' if sender == '我' else 'other'
        rows_html.append(f"""
        <tr class="msg-{is_me}">
            <td class="time">{time_str}</td>
            <td class="sender">{html.escape(sender)}</td>
            <td class="type">{msg_type}</td>
            <td class="content">{content}</td>
        </tr>""")

    html_template = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{html.escape(title)}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
            background: #f5f5f5;
            padding: 20px;
        }}
        .container {{
            max-width: 900px;
            margin: 0 auto;
            background: #fff;
            border-radius: 8px;
            box-shadow: 0 2px 12px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #07c160, #06ad56);
            color: #fff;
            padding: 20px 30px;
        }}
        .header h1 {{ font-size: 22px; margin-bottom: 5px; }}
        .header p {{ font-size: 13px; opacity: 0.9; }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th {{
            background: #fafafa;
            padding: 12px 15px;
            text-align: left;
            font-size: 13px;
            color: #666;
            border-bottom: 2px solid #e0e0e0;
        }}
        td {{
            padding: 10px 15px;
            border-bottom: 1px solid #f0f0f0;
            font-size: 14px;
            vertical-align: top;
        }}
        .msg-me {{ background: #e8f8ef; }}
        .msg-other {{ }}
        .time {{ color: #999; font-size: 12px; white-space: nowrap; width: 140px; }}
        .sender {{ font-weight: 600; color: #333; width: 100px; }}
        .type {{ color: #888; font-size: 12px; width: 70px; }}
        .content {{ word-break: break-all; line-height: 1.6; }}
        .footer {{
            text-align: center;
            padding: 15px;
            color: #999;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{html.escape(title)}</h1>
            <p>导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 共 {len(messages)} 条消息</p>
        </div>
        <table>
            <thead>
                <tr>
                    <th>时间</th>
                    <th>发送者</th>
                    <th>类型</th>
                    <th>内容</th>
                </tr>
            </thead>
            <tbody>
                {''.join(rows_html)}
            </tbody>
        </table>
        <div class="footer">
            由 WeChat Message Reader 工具生成
        </div>
    </div>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_template)

    print(f"[+] 已导出 {len(messages)} 条消息到: {output_path}")


def _sanitize_for_excel(value: str) -> str:
    """清理字符串中 Excel 不支持的非法字符"""
    if not isinstance(value, str):
        value = str(value)
    # 移除非法 XML 字符
    import re
    # 合法的 XML 1.0 字符: #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] | [#x10000-#x10FFFF]
    illegal = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
    return illegal.sub('', value)


def export_to_xlsx(messages, output_path, contacts_map=None, title=None):
    """
    导出消息到 Excel 文件 (.xlsx)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[!] 请先安装 openpyxl: pip install openpyxl")
        return False

    if contacts_map is None:
        contacts_map = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "聊天记录"

    # 样式定义
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='07C160', end_color='07C160', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0'),
    )
    cell_alignment = Alignment(vertical='top', wrap_text=True)

    # 表头
    headers = ['序号', '时间', '发送者', '消息类型', '内容']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 80

    # 写入数据
    for i, msg in enumerate(messages, 1):
        row = i + 1
        time_str = format_timestamp(msg.get('CreateTime') or msg.get('createtime'))
        sender = _sanitize_for_excel(get_sender_name(msg, contacts_map))
        msg_type = get_msg_type_name(msg.get('Type', 0))
        content = _sanitize_for_excel(get_content(msg))

        values = [i, time_str, sender, msg_type, content]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if col == 2:
                cell.alignment = Alignment(vertical='top', horizontal='center')

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = f'A1:E{len(messages) + 1}'

    wb.save(output_path)
    print(f"[+] 已导出 {len(messages)} 条消息到: {output_path}")
    return True


def export_messages(messages, output_path, fmt='html', contacts_map=None, title=None):
    """
    统一导出接口

    fmt: 'csv', 'txt', 'html', 'xlsx'
    """
    exporters = {
        'csv': export_to_csv,
        'txt': export_to_txt,
        'html': export_to_html,
        'xlsx': export_to_xlsx,
    }

    exporter = exporters.get(fmt)
    if not exporter:
        print(f"[!] 不支持的导出格式: {fmt}, 支持: csv, txt, html, xlsx")
        return False

    kwargs = {'contacts_map': contacts_map}
    if fmt in ('html', 'xlsx') and title:
        kwargs['title'] = title

    exporter(messages, output_path, **kwargs)
    return True