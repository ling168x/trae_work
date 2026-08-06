"""导出 message_0.db 全部聊天记录到 Excel - 最终版"""
import sqlite3, os, shutil, hashlib
from datetime import datetime, timezone, timedelta

try:
    import sqlcipher3
except ImportError:
    print("需要 sqlcipher3"); exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess; subprocess.run(["python", "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

KEY = '4c64d04ecd632d43eaf1cccecd944c98388da266521f19ce855e07126f7d09e7'
DB_PATH = r'D:\wechat\xwechat_files\LING827323180_96c1\db_storage\message\message_0.db'
EXPORT_DIR = r'd:\traework\wechat_msg_reader\exports'
os.makedirs(EXPORT_DIR, exist_ok=True)

tmp_path = DB_PATH + '.export3'
shutil.copy2(DB_PATH, tmp_path)
db = sqlcipher3.connect(tmp_path)
c = db.cursor()
c.execute(f"PRAGMA key=\"x'{KEY}'\"")
c.execute("PRAGMA cipher_page_size=4096")
c.execute("PRAGMA kdf_iter=256000")
c.execute("PRAGMA cipher_compatibility=4")

# Name2Id: (user_name, is_session) - hash = MD5(user_name)
c.execute("SELECT user_name FROM Name2Id")
hash_to_name = {}
for (uname,) in c.fetchall():
    if uname:
        h = hashlib.md5(uname.encode()).hexdigest()
        hash_to_name[h] = uname

print(f"Name2Id: {len(hash_to_name)} 个映射")

# 聊天列表
c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'Msg_%'")
msg_tables = [r[0] for r in c.fetchall()]

chat_info = []
for tbl in msg_tables:
    hash_id = tbl[4:]
    c.execute(f"SELECT count(*) FROM '{tbl}'")
    count = c.fetchone()[0]
    username = hash_to_name.get(hash_id, f'未知({hash_id[:8]})')
    chat_info.append((tbl, hash_id, username, count))

chat_info.sort(key=lambda x: -x[3])
for i, (tbl, h, name, cnt) in enumerate(chat_info):
    print(f"  {i+1}. [{cnt:5d}条] {name}")

# 消息类型
MSG_TYPE = {1: '文本', 3: '图片', 34: '语音', 43: '视频', 47: '表情', 49: '链接/文件', 50: '系统消息', 10000: '系统通知', 10002: '系统消息'}

TZ = timezone(timedelta(hours=8))
def ts_to_str(ts):
    try:
        return datetime.fromtimestamp(int(ts), tz=TZ).strftime('%Y-%m-%d %H:%M:%S')
    except:
        return str(ts)

def parse_content(content, msg_type):
    if content is None:
        return f'[{MSG_TYPE.get(msg_type, f"类型{msg_type}")}]', ''
    if isinstance(content, bytes):
        try:
            text = content.decode('utf-8')
            if '\n' in text:
                sender, msg = text.split('\n', 1)
                return msg, sender
            return text, ''
        except:
            return f'[{MSG_TYPE.get(msg_type, f"类型{msg_type}")}]', ''
    if '\n' in content:
        sender, msg = content.split('\n', 1)
        return msg, sender
    return content, ''

# Excel 样式
hfont = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
calign = Alignment(vertical='top', wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
headers = ['序号', '时间', '发送者', '聊天', '类型', '内容']

wb = Workbook()
ws = wb.active
ws.title = "汇总"
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = border

cur_row = 2
msg_counter = 0

for tbl, hash_id, username, count in chat_info:
    if count == 0:
        continue
    print(f"  导出 [{count}条] {username}...")
    try:
        c.execute(f"SELECT local_id, create_time, real_sender_id, message_content, local_type FROM '{tbl}' ORDER BY create_time")
        for row in c.fetchall():
            local_id, create_time, real_sender_id, content, msg_type = row
            time_str = ts_to_str(create_time) if create_time else ''
            msg_text, content_sender = parse_content(content, msg_type)
            sender = content_sender if content_sender else str(real_sender_id) if real_sender_id else ''
            type_str = MSG_TYPE.get(msg_type, f'类型{msg_type}')
            msg_counter += 1
            row_data = [msg_counter, time_str, sender, username, type_str, msg_text]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=cur_row, column=col, value=val)
                cell.alignment = calign; cell.border = border
            cur_row += 1
    except Exception as e:
        print(f"    [ERROR] {e}")

# 列宽
for col_letter, w in {'A': 8, 'B': 20, 'C': 25, 'D': 30, 'E': 12, 'F': 60}.items():
    ws.column_dimensions[col_letter].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output = os.path.join(EXPORT_DIR, f'wechat_all_{timestamp}.xlsx')
wb.save(output)
print(f"\n导出完成: {output}")
print(f"总计: {msg_counter} 条消息, {len(chat_info)} 个聊天")

db.close()
try:
    os.remove(tmp_path)
except:
    pass