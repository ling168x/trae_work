"""导出 message_0.db 中所有聊天记录到 Excel - 修复版"""
import sqlite3, os, shutil
from datetime import datetime, timezone, timedelta

try:
    import sqlcipher3
except ImportError:
    print("需要 sqlcipher3"); exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess; subprocess.run(["python", "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

KEY = '4c64d04ecd632d43eaf1cccecd944c98388da266521f19ce855e07126f7d09e7'
DB_PATH = r'D:\wechat\xwechat_files\LING827323180_96c1\db_storage\message\message_0.db'
EXPORT_DIR = r'd:\traework\wechat_msg_reader\exports'
os.makedirs(EXPORT_DIR, exist_ok=True)

# 解密数据库
tmp_path = DB_PATH + '.export2'
shutil.copy2(DB_PATH, tmp_path)
db = sqlcipher3.connect(tmp_path)
c = db.cursor()
c.execute(f"PRAGMA key=\"x'{KEY}'\"")
c.execute("PRAGMA cipher_page_size=4096")
c.execute("PRAGMA kdf_iter=256000")
c.execute("PRAGMA cipher_compatibility=4")

# 读取 Name2Id 映射: (username, hash_id)
c.execute("SELECT * FROM Name2Id")
name_to_hash = {}  # username -> hash
hash_to_name = {}  # hash -> username
for row in c.fetchall():
    username, hash_id = row[0], row[1]
    name_to_hash[username] = hash_id
    hash_to_name[hash_id] = username

print(f"Name2Id: {len(name_to_hash)} 个映射")

# 列出所有聊天
c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'Msg_%'")
msg_tables = [r[0] for r in c.fetchall()]

# 统计每个聊天
chat_info = []
for tbl in msg_tables:
    hash_id = tbl[4:]
    c.execute(f"SELECT count(*) FROM '{tbl}'")
    count = c.fetchone()[0]
    username = hash_to_name.get(hash_id, f'未知({hash_id[:8]})')
    chat_info.append((tbl, hash_id, username, count))

chat_info.sort(key=lambda x: -x[3])
print(f"聊天: {len(chat_info)} 个")

# 消息类型
MSG_TYPE = {1: '文本', 3: '图片', 34: '语音', 43: '视频', 47: '表情', 49: '链接/文件', 50: '系统消息', 10000: '系统通知', 10002: '系统消息'}

# 时间转换
TZ = timezone(timedelta(hours=8))
def ts_to_str(ts):
    try:
        return datetime.fromtimestamp(int(ts), tz=TZ).strftime('%Y-%m-%d %H:%M:%S')
    except:
        return str(ts)

# 解析消息内容
def parse_content(content, msg_type):
    if content is None:
        return f'[{MSG_TYPE.get(msg_type, f"类型{msg_type}")}]', ''
    if isinstance(content, bytes):
        # 尝试解码
        try:
            text = content.decode('utf-8')
            if '\n' in text:
                sender, msg = text.split('\n', 1)
                return msg, sender
            return text, ''
        except:
            return f'[{MSG_TYPE.get(msg_type, f"类型{msg_type}")}]', ''
    # 字符串
    if '\n' in content:
        sender, msg = content.split('\n', 1)
        return msg, sender
    return content, ''

# 样式
hfont = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
calign = Alignment(vertical='top', wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ['序号', '时间', '发送者', '聊天', '类型', '内容']

wb = Workbook()
ws = wb.active
ws.title = "汇总"
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = border

cur_row = 2
msg_counter = 0

# 导出每个聊天
for tbl, hash_id, username, count in chat_info:
    if count == 0:
        continue
    
    print(f"  导出 [{count}条] {username}...")
    
    try:
        c.execute(f"SELECT local_id, create_time, real_sender_id, message_content, local_type FROM '{tbl}' ORDER BY create_time")
        for row in c.fetchall():
            local_id, create_time, real_sender_id, content, msg_type = row
            
            # 时间
            time_str = ts_to_str(create_time) if create_time else ''
            
            # 内容
            msg_text, content_sender = parse_content(content, msg_type)
            sender = content_sender if content_sender else str(real_sender_id) if real_sender_id else ''
            
            type_str = MSG_TYPE.get(msg_type, f'类型{msg_type}')
            
            msg_counter += 1
            row_data = [msg_counter, time_str, sender, username, type_str, msg_text]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=cur_row, column=col, value=val)
                cell.alignment = calign; cell.border = border
            cur_row += 1
    except Exception as e:
        print(f"    [ERROR] {username}: {e}")

# 列宽
widths = {'A': 8, 'B': 20, 'C': 25, 'D': 30, 'E': 12, 'F': 60}
for col_letter, w in widths.items():
    ws.column_dimensions[col_letter].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output = os.path.join(EXPORT_DIR, f'wechat_all_{timestamp}.xlsx')
wb.save(output)
print(f"\n导出完成: {output}")
print(f"总计: {msg_counter} 条消息, {len(chat_info)} 个聊天")

db.close()
try:
    os.remove(tmp_path)
except:
    pass