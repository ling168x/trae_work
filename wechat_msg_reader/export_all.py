"""导出 message_0.db 中所有聊天记录到 Excel"""
import sqlite3, os, hashlib, struct, json, shutil
from datetime import datetime, timezone, timedelta

try:
    import sqlcipher3
except ImportError:
    print("需要 sqlcipher3")
    exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("安装 openpyxl...")
    import subprocess
    subprocess.run(["python", "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

KEY = '4c64d04ecd632d43eaf1cccecd944c98388da266521f19ce855e07126f7d09e7'
DB_PATH = r'D:\wechat\xwechat_files\LING827323180_96c1\db_storage\message\message_0.db'
EXPORT_DIR = r'd:\traework\wechat_msg_reader\exports'
os.makedirs(EXPORT_DIR, exist_ok=True)

# 复制并解密
tmp_path = DB_PATH + '.export'
shutil.copy2(DB_PATH, tmp_path)
db = sqlcipher3.connect(tmp_path)
c = db.cursor()
c.execute(f"PRAGMA key=\"x'{KEY}'\"")
c.execute("PRAGMA cipher_page_size=4096")
c.execute("PRAGMA kdf_iter=256000")
c.execute("PRAGMA cipher_compatibility=4")

# 列出所有聊天表
c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'Msg_%'")
msg_tables = [r[0] for r in c.fetchall()]
print(f"找到 {len(msg_tables)} 个聊天表")

# 读取 Name2Id 映射
c.execute("SELECT * FROM Name2Id")
name_map = {}
for row in c.fetchall():
    if len(row) >= 2:
        name_map[row[1]] = row[0]  # hash -> name

# 获取每个聊天的消息数量
chat_info = []
for tbl in msg_tables:
    hash_id = tbl[4:]  # 去掉 "Msg_" 前缀
    c.execute(f"SELECT count(*) FROM '{tbl}'")
    count = c.fetchone()[0]
    username = name_map.get(hash_id, '未知')
    chat_info.append((tbl, hash_id, username, count))

chat_info.sort(key=lambda x: -x[3])  # 按消息数降序
print(f"\n聊天列表 (按消息数排序):")
for i, (tbl, hash_id, username, count) in enumerate(chat_info):
    print(f"  {i+1}. [{count}条] {username} ({tbl})")

# 消息类型映射
MSG_TYPE_MAP = {
    1: '文本', 3: '图片', 34: '语音', 43: '视频',
    47: '表情', 49: '链接/小程序/文件', 50: '系统消息',
    10000: '系统通知', 10002: '系统消息',
}

# 时间戳基准 (新版微信使用自定义时间戳)
# 尝试从 TimeStamp 表获取时间基准
c.execute("SELECT * FROM TimeStamp")
time_rows = c.fetchall()
if time_rows:
    print(f"\nTimeStamp 表: {len(time_rows)} 行")
    for row in time_rows[:3]:
        print(f"  {row}")

# 导出所有聊天到一个 Excel
wb = Workbook()
ws = wb.active
ws.title = "汇总"

# 样式
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
cell_alignment = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 创建汇总表
ws_headers = ['序号', '时间', '发送者', '聊天', '消息类型', '消息内容']
for col, header in enumerate(ws_headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

cur_row = 2
msg_counter = 0

# 导出每个聊天
for tbl, hash_id, username, count in chat_info:
    if count == 0:
        continue
    
    # 为每个聊天创建独立 sheet (如果消息数 > 100)
    if count > 100:
        chat_sheet = wb.create_sheet(title=username[:31])
        chat_sheet_row = 1
        for col, header in enumerate(ws_headers, 1):
            cell = chat_sheet.cell(row=chat_sheet_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        chat_sheet_row = 2
    
    # 读取消息
    try:
        c.execute(f"SELECT local_id, server_id, create_time, real_sender_id, message_content, local_type FROM '{tbl}' ORDER BY create_time")
        for row in c.fetchall():
            local_id, server_id, create_time, real_sender_id, content, msg_type = row
            
            # 解析时间
            if create_time and create_time > 1000000000:
                # 可能是毫秒时间戳
                if create_time > 10000000000000:
                    ts = create_time / 1000
                elif create_time > 1000000000000:
                    ts = create_time / 1000
                else:
                    ts = create_time
                try:
                    dt = datetime.fromtimestamp(ts, tz=timezone(timedelta(hours=8)))
                    time_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    time_str = str(create_time)
            else:
                time_str = str(create_time)
            
            # 解析内容
            if content:
                # 格式: username:\n内容
                content_str = content
                if isinstance(content_str, bytes):
                    content_str = content_str.decode('utf-8', errors='replace')
            else:
                content_str = f'[{MSG_TYPE_MAP.get(msg_type, f"类型{msg_type}")}]'
            
            # 发送者
            sender = real_sender_id or '未知'
            
            msg_type_str = MSG_TYPE_MAP.get(msg_type, f'类型{msg_type}')
            chat_name = username
            
            # 写入汇总表
            msg_counter += 1
            row_data = [msg_counter, time_str, sender, chat_name, msg_type_str, content_str]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=cur_row, column=col, value=val)
                cell.alignment = cell_alignment
                cell.border = thin_border
            cur_row += 1
            
            # 写入独立 sheet
            if count > 100:
                row_data2 = [msg_counter, time_str, sender, chat_name, msg_type_str, content_str]
                for col, val in enumerate(row_data2, 1):
                    cell = chat_sheet.cell(row=chat_sheet_row, column=col, value=val)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                chat_sheet_row += 1
        
        print(f"  [{msg_counter}条] {username} 已导出")
    except Exception as e:
        print(f"  [ERROR] {username}: {e}")

# 设置列宽
for ws_sheet in [ws] + [wb[s] for s in wb.sheetnames if s != '汇总']:
    ws_sheet.column_dimensions['A'].width = 8
    ws_sheet.column_dimensions['B'].width = 20
    ws_sheet.column_dimensions['C'].width = 25
    ws_sheet.column_dimensions['D'].width = 30
    ws_sheet.column_dimensions['E'].width = 12
    ws_sheet.column_dimensions['F'].width = 60
    ws_sheet.freeze_panes = 'A2'
    ws_sheet.auto_filter.ref = ws_sheet.dimensions

# 保存
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output = os.path.join(EXPORT_DIR, f'wechat_all_messages_{timestamp}.xlsx')
wb.save(output)
print(f"\n导出完成: {output}")
print(f"总计: {msg_counter} 条消息, {len(chat_info)} 个聊天")

db.close()
try:
    os.remove(tmp_path)
except:
    pass