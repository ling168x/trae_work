"""解密所有数据库并导出全部聊天记录"""
import sqlite3, os, shutil, hashlib, json, sys
from datetime import datetime, timezone, timedelta

try:
    import sqlcipher3
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "sqlcipher3"])
    import sqlcipher3

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess; subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 加载密钥
KEY_CONFIG = r'd:\traework\wechat_msg_reader\extracted_keys.json'
DB_DIR = r'D:\wechat\xwechat_files\LING827323180_96c1\db_storage'
EXPORT_DIR = r'd:\traework\wechat_msg_reader\exports'
DECRYPT_DIR = r'd:\traework\wechat_msg_reader\decrypted'
os.makedirs(EXPORT_DIR, exist_ok=True)
os.makedirs(DECRYPT_DIR, exist_ok=True)

# 消息类型
MSG_TYPE = {1: '文本', 3: '图片', 34: '语音', 43: '视频', 47: '表情', 49: '链接/文件', 50: '系统消息', 10000: '系统通知', 10002: '系统消息'}
TZ = timezone(timedelta(hours=8))

def ts_to_str(ts):
    try:
        return datetime.fromtimestamp(int(ts), tz=TZ).strftime('%Y-%m-%d %H:%M:%S')
    except:
        return str(ts)

def parse_content(content, msg_type):
    if content is None:
        return f'[{MSG_TYPE.get(msg_type, f"类型{msg_type}")}]', ''
    if isinstance(content, bytes):
        try:
            text = content.decode('utf-8')
            if '\n' in text:
                sender, msg = text.split('\n', 1)
                return msg, sender
            return text, ''
        except:
            return f'[{MSG_TYPE.get(msg_type, f"类型{msg_type}")}]', ''
    if '\n' in content:
        sender, msg = content.split('\n', 1)
        return msg, sender
    return content, ''

def decrypt_and_verify(db_path, key):
    """解密数据库并验证"""
    if not os.path.exists(db_path):
        return None
    try:
        tmp = db_path + '.decrypt'
        shutil.copy2(db_path, tmp)
        db = sqlcipher3.connect(tmp)
        c = db.cursor()
        c.execute(f"PRAGMA key=\"x'{key}'\"")
        c.execute("PRAGMA cipher_page_size=4096")
        c.execute("PRAGMA kdf_iter=256000")
        c.execute("PRAGMA cipher_compatibility=4")
        c.execute("SELECT count(*) FROM sqlite_master")
        c.fetchone()
        return db, tmp
    except Exception as e:
        try:
            db.close()
        except:
            pass
        return None

def export_messages(db, c, output_path):
    """导出消息到 Excel"""
    # 读取 Name2Id
    c.execute("SELECT user_name FROM Name2Id")
    hash_to_name = {}
    for (uname,) in c.fetchall():
        if uname:
            h = hashlib.md5(uname.encode()).hexdigest()
            hash_to_name[h] = uname

    # 读取聊天表
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'Msg_%'")
    tables = [r[0] for r in c.fetchall()]

    if not tables:
        print("  没有找到聊天表")
        return False

    # 统计
    chat_info = []
    for tbl in tables:
        hash_id = tbl[4:]
        c.execute(f"SELECT count(*) FROM '{tbl}'")
        count = c.fetchone()[0]
        username = hash_to_name.get(hash_id, f'未知({hash_id[:8]})')
        chat_info.append((tbl, hash_id, username, count))

    chat_info.sort(key=lambda x: -x[3])

    # Excel 样式
    hfont = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    calign = Alignment(vertical='top', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ['序号', '时间', '发送者', '聊天', '类型', '内容']

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = border

    cur_row = 2
    msg_counter = 0

    for tbl, hash_id, username, count in chat_info:
        if count == 0:
            continue
        print(f"  [{count}条] {username}...")
        try:
            c.execute(f"SELECT local_id, create_time, real_sender_id, message_content, local_type FROM '{tbl}' ORDER BY create_time")
            for row in c.fetchall():
                local_id, create_time, real_sender_id, content, msg_type = row
                time_str = ts_to_str(create_time) if create_time else ''
                msg_text, content_sender = parse_content(content, msg_type)
                sender = content_sender if content_sender else str(real_sender_id) if real_sender_id else ''
                msg_counter += 1
                row_data = [msg_counter, time_str, sender, username, MSG_TYPE.get(msg_type, f'类型{msg_type}'), msg_text]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=cur_row, column=col, value=val)
                    cell.alignment = calign; cell.border = border
                cur_row += 1
        except Exception as e:
            print(f"    [ERROR] {e}")

    for col_letter, w in {'A': 8, 'B': 20, 'C': 25, 'D': 30, 'E': 12, 'F': 60}.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"  导出完成: {output_path}")
    print(f"  总计: {msg_counter} 条消息, {len(chat_info)} 个聊天")
    return True

def main():
    print("=" * 60)
    print("  微信数据库解密 & 导出工具")
    print("=" * 60)

    # 加载密钥
    if not os.path.exists(KEY_CONFIG):
        print(f"\n[错误] 未找到密钥文件: {KEY_CONFIG}")
        print("  请先运行 extract_keys.py 提取密钥")
        return
    with open(KEY_CONFIG, 'r') as f:
        config = json.load(f)

    key = config.get('key')
    if not key:
        print("[错误] 密钥文件中没有 key 字段")
        return

    print(f"\n密钥: {key[:16]}...")

    # 扫描所有数据库
    print("\n扫描数据库文件...")
    dbs = []
    for root, dirs, files in os.walk(DB_DIR):
        for f in files:
            if f.endswith('.db') and '-wal' not in f and '-shm' not in f:
                dbs.append(os.path.join(root, f))

    print(f"找到 {len(dbs)} 个数据库")

    # 解密每个数据库
    print("\n解密数据库...")
    decrypted = {}
    for db_path in dbs:
        rel = os.path.relpath(db_path, DB_DIR)
        result = decrypt_and_verify(db_path, key)
        if result:
            db, tmp = result
            print(f"  [OK] {rel}")
            decrypted[rel] = (db, tmp)
        else:
            print(f"  [MISS] {rel}")

    print(f"\n成功解密: {len(decrypted)}/{len(dbs)} 个数据库")

    # 导出消息
    print("\n导出消息...")
    for rel in ['message\\message_0.db', 'message\\message_1.db', 'message\\message_2.db']:
        if rel in decrypted:
            db, tmp = decrypted[rel]
            output = os.path.join(EXPORT_DIR, f'wechat_{rel.replace(chr(92), "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
            print(f"\n处理: {rel}")
            export_messages(db, db.cursor(), output)
            db.close()
            try:
                os.remove(tmp)
            except:
                pass

    # 清理
    for db, tmp in decrypted.values():
        try:
            db.close()
        except:
            pass
        try:
            os.remove(tmp)
        except:
            pass

    print("\n完成！")

if __name__ == '__main__':
    main()