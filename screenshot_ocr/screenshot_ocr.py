"""
截图提取文字工具
- 微信风格截图：屏幕实景 + 暗色遮罩，选区清晰可见
- 自动 OCR 识别截图中的文字（中英文）
- 每次识别结果追加到同一个 Excel 文件中
"""

import os
import sys

# PaddlePaddle 3.x 在 Windows + OneDNN 下存在 PIR 执行器 bug：
#   NotImplementedError: ConvertPirAttribute2RuntimeAttribute not support
#   [pir::ArrayAttribute<pir::DoubleAttribute>]
# 必须在 import paddle / paddleocr 之前关闭 PIR API 与 mkldnn。
os.environ.setdefault("FLAGS_enable_pir_api", "0")
os.environ.setdefault("FLAGS_enable_pir_in_executor", "0")
os.environ.setdefault("FLAGS_use_mkldnn", "0")

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import traceback
from datetime import datetime

from PIL import Image, ImageGrab, ImageTk, ImageEnhance, ImageFilter
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ==================== 配置 ====================
DEFAULT_EXCEL_NAME = "screenshot_texts.xlsx"

_ocr = None
_ocr_lock = threading.Lock()


def get_ocr():
    global _ocr
    if _ocr is None:
        with _ocr_lock:
            if _ocr is None:
                from paddleocr import PaddleOCR
                _ocr = PaddleOCR(
                    use_angle_cls=True, lang="ch",
                    enable_mkldnn=False,
                )
    return _ocr


# ==================== 微信风格截图覆盖层 ====================
class ScreenshotOverlay:
    """屏幕实景 + 暗色遮罩，选区清晰可见"""

    def __init__(self, on_complete):
        self.on_complete = on_complete
        self.start_x = None
        self.start_y = None
        self.selection_rect = None

        # 先截取全屏作为背景
        self.screenshot_img = ImageGrab.grab()
        self.screenshot_tk = ImageTk.PhotoImage(self.screenshot_img)

        self.top = tk.Toplevel()
        self.top.attributes("-fullscreen", True)
        self.top.attributes("-topmost", True)
        self.top.configure(bg="black")

        self.sw = self.top.winfo_screenwidth()
        self.sh = self.top.winfo_screenheight()

        self.canvas = tk.Canvas(
            self.top, cursor="cross",
            width=self.sw, height=self.sh,
            highlightthickness=0, bg="black",
        )
        self.canvas.pack()

        # 屏幕截图作为背景
        self.canvas.create_image(
            0, 0, image=self.screenshot_tk, anchor="nw", tag="bg",
        )

        # 初始全屏暗色遮罩
        self._draw_dim(0, 0, self.sw, self.sh)

        # 提示文字
        self.canvas.create_text(
            self.sw // 2, 30,
            text="拖动鼠标框选截图区域    ESC 取消",
            fill="white",
            font=("微软雅黑", 13, "bold"),
            tag="hint",
        )

        self.canvas.bind("<ButtonPress-1>", self._on_down)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_up)
        self.top.bind("<Escape>", self._cancel)

    def _draw_dim(self, x1, y1, x2, y2):
        self.canvas.delete("dim")
        self.canvas.create_rectangle(
            x1, y1, x2, y2,
            fill="black", stipple="gray50",
            outline="", tag="dim",
        )

    def _update_dim_around_selection(self, sx1, sy1, sx2, sy2):
        self.canvas.delete("dim")
        if sy1 > 0:
            self.canvas.create_rectangle(
                0, 0, self.sw, sy1,
                fill="black", stipple="gray50", outline="", tag="dim",
            )
        if sy2 < self.sh:
            self.canvas.create_rectangle(
                0, sy2, self.sw, self.sh,
                fill="black", stipple="gray50", outline="", tag="dim",
            )
        if sx1 > 0:
            self.canvas.create_rectangle(
                0, sy1, sx1, sy2,
                fill="black", stipple="gray50", outline="", tag="dim",
            )
        if sx2 < self.sw:
            self.canvas.create_rectangle(
                sx2, sy1, self.sw, sy2,
                fill="black", stipple="gray50", outline="", tag="dim",
            )

    def _on_down(self, event):
        self.start_x = event.x
        self.start_y = event.y
        self.canvas.delete("selection", "hint", "size_label")
        self.selection_rect = self.canvas.create_rectangle(
            self.start_x, self.start_y, self.start_x, self.start_y,
            outline="#00C853", width=2, tag="selection",
        )

    def _on_drag(self, event):
        if self.selection_rect is None:
            return
        x1 = min(self.start_x, event.x)
        y1 = min(self.start_y, event.y)
        x2 = max(self.start_x, event.x)
        y2 = max(self.start_y, event.y)

        self.canvas.coords(self.selection_rect, x1, y1, x2, y2)
        self._update_dim_around_selection(x1, y1, x2, y2)

        self.canvas.delete("size_label")
        w, h = x2 - x1, y2 - y1
        label_x = x1 + 5
        label_y = y1 - 5 if y1 > 25 else y1 + 20
        self.canvas.create_text(
            label_x, label_y,
            text=f"{w} x {h}",
            fill="#00C853", anchor="sw",
            font=("微软雅黑", 9, "bold"),
            tag="size_label",
        )

    def _on_up(self, event):
        if self.start_x is None:
            return
        x1 = min(self.start_x, event.x)
        y1 = min(self.start_y, event.y)
        x2 = max(self.start_x, event.x)
        y2 = max(self.start_y, event.y)
        self.top.destroy()
        if x2 - x1 >= 10 and y2 - y1 >= 10:
            self.on_complete(x1, y1, x2, y2)
        else:
            self.on_complete(None, None, None, None)

    def _cancel(self, event=None):
        self.top.destroy()
        self.on_complete(None, None, None, None)


# ==================== 主应用 ====================
class ScreenshotOCRApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("截图提取文字工具")
        self.root.geometry("460x520")
        self.root.minsize(420, 420)
        self.root.configure(bg="#f5f5f5")

        self.excel_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), DEFAULT_EXCEL_NAME
        )

        self._setup_ui()
        self._setup_hotkey()
        self._update_excel_info()

    # ---------- UI ----------
    def _setup_ui(self):
        header = tk.Frame(self.root, bg="#2c3e50", height=56)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(
            header, text="截图提取文字工具",
            font=("微软雅黑", 15, "bold"), fg="white", bg="#2c3e50",
        ).pack(expand=True)

        btn_frame = tk.Frame(self.root, bg="#f5f5f5")
        btn_frame.pack(pady=14)

        self.btn_capture = tk.Button(
            btn_frame,
            text="开始截图  (Ctrl+Shift+S)",
            command=self._start_screenshot,
            font=("微软雅黑", 12, "bold"),
            width=28, height=2,
            bg="#27ae60", fg="white",
            activebackground="#219a52", activeforeground="white",
            border=0, cursor="hand2",
        )
        self.btn_capture.pack()

        path_frame = tk.LabelFrame(
            self.root, text="Excel 保存路径", bg="#f5f5f5", font=("微软雅黑", 9),
        )
        path_frame.pack(fill=tk.X, padx=20, pady=(5, 8))

        inner = tk.Frame(path_frame, bg="#f5f5f5")
        inner.pack(fill=tk.X, padx=8, pady=8)

        self.path_var = tk.StringVar(value=self.excel_path)
        self.path_entry = tk.Entry(
            inner, textvariable=self.path_var,
            font=("微软雅黑", 9), state="readonly",
        )
        self.path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        tk.Button(
            inner, text="更改", command=self._browse_path,
            font=("微软雅黑", 9), width=5, bg="#3498db", fg="white",
            border=0, cursor="hand2",
        ).pack(side=tk.RIGHT)

        self.status_var = tk.StringVar(value="就绪")
        tk.Label(
            self.root, textvariable=self.status_var,
            font=("微软雅黑", 9), fg="#7f8c8d", bg="#f5f5f5", anchor="w",
        ).pack(fill=tk.X, padx=20, pady=(0, 3))

        self.excel_info_var = tk.StringVar(value="")
        tk.Label(
            self.root, textvariable=self.excel_info_var,
            font=("微软雅黑", 8), fg="#95a5a6", bg="#f5f5f5", anchor="w",
        ).pack(fill=tk.X, padx=20, pady=(0, 8))

        result_frame = tk.LabelFrame(
            self.root, text="最近识别结果", bg="#f5f5f5", font=("微软雅黑", 9),
        )
        result_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 12))

        self.result_text = tk.Text(
            result_frame, font=("微软雅黑", 10), wrap=tk.WORD,
            bg="white", relief=tk.FLAT, border=2, padx=8, pady=8,
        )
        self.result_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        bottom = tk.Frame(self.root, bg="#f5f5f5")
        bottom.pack(fill=tk.X, padx=20, pady=(0, 12))
        tk.Button(
            bottom, text="打开 Excel 文件", command=self._open_excel,
            font=("微软雅黑", 9), bg="#8e44ad", fg="white",
            border=0, cursor="hand2", padx=12, pady=4,
        ).pack(side=tk.LEFT)
        tk.Button(
            bottom, text="清除所有记录", command=self._clear_records,
            font=("微软雅黑", 9), bg="#e74c3c", fg="white",
            border=0, cursor="hand2", padx=12, pady=4,
        ).pack(side=tk.RIGHT)

    def _setup_hotkey(self):
        try:
            import keyboard
            # 通过 root.after 回到主线程执行，避免 tkinter 线程安全问题
            keyboard.add_hotkey(
                "ctrl+shift+s",
                lambda: self.root.after(0, self._start_screenshot),
            )
            self.status_var.set("就绪 — Ctrl+Shift+S 截图")
        except Exception:
            self.status_var.set("就绪 — 点击按钮截图")

    # ---------- 截图流程 ----------
    def _start_screenshot(self):
        self.root.iconify()
        self.root.after(300, lambda: ScreenshotOverlay(on_complete=self._on_capture_done))

    def _on_capture_done(self, x1, y1, x2, y2):
        self.root.deiconify()
        if x1 is None:
            self.status_var.set("已取消")
            return

        self.status_var.set("截图完成，识别中...")
        self.btn_capture.config(state=tk.DISABLED, text="处理中...")

        thread = threading.Thread(
            target=self._process, args=(x1, y1, x2, y2), daemon=True,
        )
        thread.start()

    def _preprocess_image(self, img):
        """图像预处理：放大、增强对比度，提升 OCR 识别率"""
        w, h = img.size

        # 如果截图太小，放大 2~3 倍
        if w < 400 or h < 200:
            scale = 3 if w < 200 else 2
            img = img.resize((w * scale, h * scale), Image.LANCZOS)

        # 增强对比度
        img = ImageEnhance.Contrast(img).enhance(1.8)

        # 锐化
        img = img.filter(ImageFilter.SHARPEN)

        return img

    def _process(self, x1, y1, x2, y2):
        try:
            img = ImageGrab.grab(bbox=(x1, y1, x2, y2))

            self._update_status("正在预处理图像...")
            img = self._preprocess_image(img)

            self._update_status("正在加载 OCR 模型...")
            ocr = get_ocr()

            self._update_status("正在 OCR 识别...")
            import numpy as np
            img_np = np.array(img)

            # 兼容 PaddleOCR 2.x / 3.x：
            # 3.x 用 predict()，返回 list[dict]，含 rec_texts 字段；
            # 2.x 用 ocr(img, cls=True)，返回嵌套 list。
            if hasattr(ocr, "predict"):
                results = ocr.predict(img_np)
                lines = []
                for r in results or []:
                    texts = (
                        r.get("rec_texts", [])
                        if isinstance(r, dict)
                        else getattr(r, "rec_texts", [])
                    )
                    if texts:
                        lines.extend(texts)
                    elif isinstance(r, dict) and r.get("rec_text"):
                        lines.append(r["rec_text"])
                text = "\n".join(lines).strip()
            else:
                results = ocr.ocr(img_np, cls=True)
                if results and results[0]:
                    lines = [item[1][0] for item in results[0]]
                    text = "\n".join(lines).strip()
                else:
                    text = ""

            if text:
                self._save_to_excel(text)
                self._show_result(text)
                self._update_status("识别成功！已保存到 Excel")
                self._update_excel_info()
            else:
                self._show_result("(未识别到文字，请尝试选择更清晰的区域)")
                self._update_status("未识别到文字")
        except Exception as e:
            err_msg = traceback.format_exc()
            self._show_result(f"识别出错：\n{err_msg}")
            self._update_status(f"识别失败：{e}")
        finally:
            self.root.after(0, lambda: self.btn_capture.config(
                state=tk.NORMAL, text="开始截图  (Ctrl+Shift+S)",
            ))

    def _save_to_excel(self, text):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if os.path.exists(self.excel_path):
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "截图文字记录"
            header_fill = PatternFill(
                start_color="002c3e50", end_color="002c3e50", fill_type="solid",
            )
            header_font = Font(name="微软雅黑", size=10, bold=True, color="00FFFFFF")
            for col, title in enumerate(["序号", "截图时间", "提取的文字"], 1):
                cell = ws.cell(row=1, column=col, value=title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 60

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=next_row - 1)
        ws.cell(row=next_row, column=2, value=timestamp)
        ws.cell(row=next_row, column=3, value=text)

        for col in range(1, 4):
            cell = ws.cell(row=next_row, column=col)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col <= 2:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        wb.save(self.excel_path)

    # ---------- UI 辅助 ----------
    def _show_result(self, text):
        self.root.after(0, lambda: self._set_result(text))

    def _set_result(self, text):
        self.result_text.delete("1.0", tk.END)
        self.result_text.insert("1.0", text)

    def _update_status(self, msg):
        self.root.after(0, lambda: self.status_var.set(msg))

    def _update_excel_info(self):
        count = 0
        if os.path.exists(self.excel_path):
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                ws = wb.active
                count = max(0, ws.max_row - 1)
                wb.close()
            except Exception:
                pass
        self.excel_info_var.set(f"已记录 {count} 条  |  {self.excel_path}")

    def _browse_path(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=DEFAULT_EXCEL_NAME,
            initialdir=os.path.dirname(self.excel_path),
        )
        if path:
            self.excel_path = path
            self.path_var.set(path)
            self._update_excel_info()

    def _open_excel(self):
        if os.path.exists(self.excel_path):
            os.startfile(self.excel_path)
        else:
            messagebox.showinfo("提示", "还没有记录，请先截图提取文字")

    def _clear_records(self):
        if not os.path.exists(self.excel_path):
            return
        if messagebox.askyesno("确认清除", "确定要清除所有截图文字记录吗？此操作不可恢复。"):
            try:
                os.remove(self.excel_path)
                self._update_excel_info()
                self.status_var.set("已清除所有记录")
            except Exception as e:
                messagebox.showerror("错误", f"清除失败：{e}")

    def run(self):
        self.root.mainloop()


# ==================== 入口 ====================
if __name__ == "__main__":
    app = ScreenshotOCRApp()
    app.run()